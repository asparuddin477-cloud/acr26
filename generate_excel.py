import json
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

seed_path = os.path.join('data', 'legacy_seed.json')
out_xlsx_path = os.path.join('data', 'database_peserta_acr2026.xlsx')

with open(seed_path, 'r', encoding='utf-8') as f:
    data = json.load(f)

peserta_list = data.get('peserta', [])

cols = [
    'kode', 'nama', 'ktp', 'gender', 'kategori', 'bibName', 'wa', 'jersey',
    'komunitas', 'alamat', 'provinsi', 'kota', 'darurat', 'komorbid',
    'tagihan', 'diskon', 'status', 'bibNumber', 'checkedIn',
    'kodeLogistik', 'logistikDiambil', 'bukti', 'createdAt'
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Database Peserta ACR 2026"

# Header styling
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark Blue
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

ws.row_dimensions[1].height = 28

for col_idx, col_name in enumerate(cols, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Data rows
for row_idx, p in enumerate(peserta_list, start=2):
    ws.row_dimensions[row_idx].height = 20
    is_even = (row_idx % 2 == 0)
    row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if is_even else None
    
    for col_idx, col_name in enumerate(cols, start=1):
        raw_val = p.get(col_name, '')
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        cell.font = Font(name="Calibri", size=10)

        if col_name == 'ktp':
            cell.number_format = '@'  # Text format
            cell.value = str(raw_val).strip() if raw_val else ''
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_name in ['wa', 'darurat']:
            cell.number_format = '@'
            val_str = str(raw_val).strip()
            if val_str and not val_str.startswith('0') and not val_str.startswith('+'):
                val_str = '0' + val_str
            cell.value = val_str
            cell.alignment = Alignment(horizontal="left", vertical="center")
        elif col_name == 'bukti':
            cell.value = '[Ada Bukti Transfer]' if raw_val else '[Tidak Ada Bukti]'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_name == 'checkedIn':
            cell.value = 'TRUE' if raw_val in [True, 'TRUE', 'true'] else 'FALSE'
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_name == 'createdAt':
            if isinstance(raw_val, (int, float)) and raw_val > 1000000000:
                try:
                    ts = raw_val / 1000 if raw_val > 100000000000 else raw_val
                    cell.value = datetime.fromtimestamp(ts).strftime('%Y-%m-%d %H:%M:%S')
                except Exception:
                    cell.value = str(raw_val)
            else:
                cell.value = str(raw_val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_name in ['tagihan', 'diskon']:
            try:
                cell.value = int(raw_val)
                cell.number_format = '#,##0'
            except (ValueError, TypeError):
                cell.value = raw_val
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif col_name in ['kode', 'bibNumber', 'kodeLogistik', 'gender']:
            cell.value = str(raw_val).strip() if raw_val else ''
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            if isinstance(raw_val, str):
                cell.value = raw_val.replace('\r', ' ').replace('\n', ' ').strip()
            else:
                cell.value = raw_val
            cell.alignment = Alignment(horizontal="left", vertical="center")

# Auto-adjust column widths
for col_idx in range(1, len(cols) + 1):
    col_letter = get_column_letter(col_idx)
    max_len = max(len(str(ws.cell(row=r, column=col_idx).value or '')) for r in range(1, min(len(peserta_list) + 2, 100)))
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb.save(out_xlsx_path)
print(f"Berhasil membuat file Excel (.xlsx): {out_xlsx_path} ({len(peserta_list)} baris).")
